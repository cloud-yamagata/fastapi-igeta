from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExcelExportOptions:
    sheet_name: str | None = None
    # データ書き込み開始行（marker が無い場合のフォールバック）
    start_row: int = 2
    start_col: int = 1
    # テンプレ内にこの文字列があれば、そのセル位置をデータ開始とする（セルの文字は消す）
    start_cell_marker: str | None = None
    # 見出し行を書き換えるか（タイトル領域を壊したくないため既定は False）
    write_header: bool = False
    # write_header=True の場合の見出し行（省略時は start_row-1）
    header_row: int | None = None
    # 列幅を自動調整するか（テンプレのレイアウト崩れを避けるため既定は False）
    autosize_columns: bool = False


def _find_marker_cell(ws, marker: str) -> tuple[int, int] | None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == marker:
                return (cell.row, cell.column)
    return None


def _to_excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def _inherited_number_format(ws, column: int, header_row: int | None) -> str | None:
    """列または見出しセルに設定済みの表示書式を返す（General は未設定扱い）。"""
    letter = get_column_letter(column)
    dim = ws.column_dimensions.get(letter)
    col_nf = getattr(dim, "number_format", None) if dim is not None else None
    if col_nf and col_nf != "General":
        return col_nf
    if header_row is not None and header_row > 0:
        header_nf = ws.cell(row=header_row, column=column).number_format
        if header_nf and header_nf != "General":
            return header_nf
    return None


def _set_cell_value(ws, *, row: int, column: int, value: Any, header_row: int | None) -> None:
    """値のみセットし、列に定義済みの数値書式（カンマ・小数・マイナス赤）を維持する。"""
    cell = ws.cell(row=row, column=column)
    existing_nf = cell.number_format
    cell.value = _to_excel_value(value)
    if existing_nf and existing_nf != "General":
        cell.number_format = existing_nf
        return
    inherited = _inherited_number_format(ws, column, header_row)
    if inherited:
        cell.number_format = inherited


def export_to_excel_using_template(
    *,
    template_bytes: bytes,
    rows: list[dict[str, Any]],
    columns: list[str] | None = None,
    options: ExcelExportOptions = ExcelExportOptions(),
) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[options.sheet_name] if options.sheet_name else wb.active

    if not rows:
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    if columns is None:
        # preserve order as returned by SQLAlchemy mappings
        columns = list(rows[0].keys())

    start_row = options.start_row
    start_col = options.start_col
    if options.start_cell_marker:
        hit = _find_marker_cell(ws, options.start_cell_marker)
        if hit:
            start_row, start_col = hit
            ws.cell(row=start_row, column=start_col, value=None)

    # header（既定では書かない）
    header_row_for_style = options.header_row if options.header_row is not None else (start_row - 1 if start_row > 1 else None)
    if options.write_header:
        header_row = header_row_for_style if header_row_for_style is not None else (start_row - 1)
        for i, col in enumerate(columns):
            ws.cell(row=header_row, column=start_col + i, value=col)

    # body: 値だけ流し込み、列の数値書式は維持する
    for r_i, row in enumerate(rows):
        for c_i, key in enumerate(columns):
            _set_cell_value(
                ws,
                row=start_row + r_i,
                column=start_col + c_i,
                value=row.get(key),
                header_row=header_row_for_style,
            )

    if options.autosize_columns:
        # modest autosize
        for c_i, key in enumerate(columns):
            col_letter = get_column_letter(start_col + c_i)
            ws.column_dimensions[col_letter].width = max(10, min(40, len(str(key)) + 2))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

