"""販売計画 Excel 取込（プレビュー・登録）。"""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.entities.te_monthly_product_plan.model import TeMonthlyProductPlan
from app.entities.te_monthly_product_plan.repository import TeMonthlyProductPlanRepository
from app.entities.te_monthly_sales_plan.model import TeMonthlySalesPlan
from app.entities.te_monthly_sales_plan.repository import TeMonthlySalesPlanRepository
from app.entities.tr_item.model import TrItem
from app.entities.tr_item_bom.model import TrItemBom
from app.entities.tr_sales_link_name.repository import TrSalesLinkNameRepository
from app.schemas.sales_plan_excel_import import (
    SalesPlanImportErrorEntry,
    SalesPlanImportPreviewResponse,
    SalesPlanImportProductRow,
    SalesPlanImportRegisterResponse,
    SalesPlanImportSalesRow,
    SalesPlanImportSummary,
)

SRC_COL_NAME = 2  # B
SRC_COL_KIND = 3  # C
SRC_QTY_FIRST = 4  # D
SRC_QTY_LAST = 9  # I
SRC_KIND_QTY = "予定数量"
B1_YEAR_MONTH_RE = re.compile(r"販売計画表\s*(\d{4})年(\d{1,2})月")

STATUS_OK = "ok"
STATUS_LINK_NOT_FOUND = "link_not_found"
STATUS_BOM_NOT_FOUND = "bom_not_found"


@dataclass
class BreakdownCell:
    column: str
    qty: int


@dataclass
class ParsedExcelRow:
    sales_item_name: str
    breakdowns: list[BreakdownCell] = field(default_factory=list)

    @property
    def sales_total(self) -> int:
        return sum(cell.qty for cell in self.breakdowns)


@dataclass
class BomExpansion:
    bulk_no: int
    item_name: str
    package_size: int
    need_size: int


class SalesPlanExcelImportError(ValueError):
    """取込処理の入力・解析エラー。"""


def normalize_sales_name(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "")
    text = text.replace(" ", "").replace("\u3000", "")
    text = text.replace("ティーバッグ", "TB").replace("ﾃｨｰﾊﾞｯｸﾞ", "TB")
    text = text.replace("グルメテーブル", "GT")
    return text.lower()


def _as_int_qty(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        if value <= 0:
            return None
        return int(value) if float(value).is_integer() else int(round(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    if num <= 0:
        return None
    return int(num) if float(num).is_integer() else int(round(num))


def parse_year_month_from_b1(value: Any) -> tuple[int, int]:
    text = str(value or "").strip()
    matched = B1_YEAR_MONTH_RE.search(text)
    if not matched:
        raise SalesPlanExcelImportError(
            f"B1 セルから年月を読み取れません（形式: 販売計画表　20xx年x月）: {text!r}"
        )
    year = int(matched.group(1))
    month = int(matched.group(2))
    if year < 2000 or year > 2100 or month < 1 or month > 12:
        raise SalesPlanExcelImportError(f"B1 の年月が不正です: {year}年{month}月")
    return year, month


def extract_rows_from_worksheet(ws) -> list[ParsedExcelRow]:
    rows: list[ParsedExcelRow] = []
    for row_idx in range(1, ws.max_row + 1):
        kind = ws.cell(row_idx, SRC_COL_KIND).value
        name = ws.cell(row_idx, SRC_COL_NAME).value
        if kind != SRC_KIND_QTY:
            continue
        if name is None or str(name).strip() == "":
            continue
        breakdowns: list[BreakdownCell] = []
        for col in range(SRC_QTY_FIRST, SRC_QTY_LAST + 1):
            qty = _as_int_qty(ws.cell(row_idx, col).value)
            if qty is None:
                continue
            breakdowns.append(BreakdownCell(column=get_column_letter(col), qty=qty))
        if not breakdowns:
            continue
        rows.append(ParsedExcelRow(sales_item_name=str(name).strip(), breakdowns=breakdowns))
    if not rows:
        raise SalesPlanExcelImportError("先頭シートから予定数量行を抽出できませんでした")
    return rows


def parse_sales_plan_excel(content: bytes) -> tuple[int, int, list[ParsedExcelRow]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        year, month = parse_year_month_from_b1(ws.cell(1, SRC_COL_NAME).value)
        rows = extract_rows_from_worksheet(ws)
        return year, month, rows
    finally:
        wb.close()


def build_normalized_link_lookup(exact: dict[str, int]) -> dict[str, int]:
    normalized: dict[str, int] = {}
    for name, item_no in exact.items():
        key = normalize_sales_name(name)
        if key and key not in normalized:
            normalized[key] = item_no
    return normalized


def resolve_item_no(
    sales_name: str,
    exact: dict[str, int],
    normalized: dict[str, int],
) -> int | None:
    name = sales_name.strip()
    if not name:
        return None
    if name in exact:
        return exact[name]
    key = normalize_sales_name(name)
    return normalized.get(key)


def load_item_map(session: Session) -> dict[int, TrItem]:
    rows = session.scalars(select(TrItem)).all()
    out: dict[int, TrItem] = {}
    for row in rows:
        if row.item_no is not None:
            out[int(row.item_no)] = row
    return out


def load_bom_map(session: Session) -> dict[int, TrItemBom]:
    rows = session.scalars(select(TrItemBom)).all()
    return {int(row.parent_item_no): row for row in rows}


def expand_bom(
    item_no: int,
    sales_size: int,
    item_map: dict[int, TrItem],
    bom_map: dict[int, TrItemBom],
) -> BomExpansion | None:
    parent = item_map.get(item_no)
    bom = bom_map.get(item_no)
    if parent is None or bom is None:
        return None
    child = item_map.get(int(bom.child_item_no))
    if child is None:
        return None
    package_size = int(parent.package_size or 0)
    need_size = int(round(sales_size * package_size / 1000))
    return BomExpansion(
        bulk_no=int(bom.child_item_no),
        item_name=(child.item_name or "").strip(),
        package_size=package_size,
        need_size=need_size,
    )


def build_preview(
    session: Session,
    *,
    file_name: str,
    content: bytes,
) -> SalesPlanImportPreviewResponse:
    year, month, parsed_rows = parse_sales_plan_excel(content)
    exact_links = TrSalesLinkNameRepository.build_exact_lookup(session)
    normalized_links = build_normalized_link_lookup(exact_links)
    item_map = load_item_map(session)
    bom_map = load_bom_map(session)

    sales_rows: list[SalesPlanImportSalesRow] = []
    product_rows: list[SalesPlanImportProductRow] = []

    for parsed in parsed_rows:
        item_no = resolve_item_no(parsed.sales_item_name, exact_links, normalized_links)
        if item_no is None:
            sales_rows.append(
                SalesPlanImportSalesRow(
                    sales_item_name=parsed.sales_item_name,
                    sales_size=parsed.sales_total,
                    status=STATUS_LINK_NOT_FOUND,
                    message="販売商品名リンク未登録",
                )
            )
            for cell in parsed.breakdowns:
                product_rows.append(
                    SalesPlanImportProductRow(
                        sales_item_name=parsed.sales_item_name,
                        column=cell.column,
                        qty=cell.qty,
                        status=STATUS_LINK_NOT_FOUND,
                        message="販売商品名リンク未登録",
                    )
                )
            continue

        parent = item_map.get(item_no)
        master_name = (parent.item_name if parent else "") or ""
        sales_rows.append(
            SalesPlanImportSalesRow(
                sales_item_name=parsed.sales_item_name,
                item_no=item_no,
                item_name=master_name,
                sales_size=parsed.sales_total,
                status=STATUS_OK,
            )
        )

        for cell in parsed.breakdowns:
            expanded = expand_bom(item_no, cell.qty, item_map, bom_map)
            if expanded is None:
                product_rows.append(
                    SalesPlanImportProductRow(
                        sales_item_name=parsed.sales_item_name,
                        column=cell.column,
                        qty=cell.qty,
                        item_no=item_no,
                        status=STATUS_BOM_NOT_FOUND,
                        message="商品原料対照表が未登録",
                    )
                )
                continue
            product_rows.append(
                SalesPlanImportProductRow(
                    sales_item_name=parsed.sales_item_name,
                    column=cell.column,
                    qty=cell.qty,
                    item_no=item_no,
                    bulk_no=expanded.bulk_no,
                    item_name=expanded.item_name,
                    package_size=expanded.package_size,
                    need_size=expanded.need_size,
                    status=STATUS_OK,
                )
            )

    link_not_found = sum(1 for r in sales_rows if r.status == STATUS_LINK_NOT_FOUND)
    bom_not_found = sum(1 for r in product_rows if r.status == STATUS_BOM_NOT_FOUND)
    bom_missing_labels: list[str] = []
    seen_bom: set[str] = set()
    for row in product_rows:
        if row.status != STATUS_BOM_NOT_FOUND or row.item_no is None:
            continue
        label = f"{row.item_no}: {row.sales_item_name}"
        if label not in seen_bom:
            seen_bom.add(label)
            bom_missing_labels.append(label)
    ok_sales = sum(1 for r in sales_rows if r.status == STATUS_OK)
    ok_product = sum(1 for r in product_rows if r.status == STATUS_OK)
    item_no_counts: dict[int, int] = {}
    for row in sales_rows:
        if row.status != STATUS_OK or row.item_no is None:
            continue
        item_no_counts[int(row.item_no)] = item_no_counts.get(int(row.item_no), 0) + 1
    duplicate_item_nos = [
        f"item_no={item_no}（{count}行）"
        for item_no, count in sorted(item_no_counts.items())
        if count > 1
    ]
    merged_sales_rows = len(item_no_counts)
    can_register = (
        len(sales_rows) > 0
        and link_not_found == 0
        and bom_not_found == 0
        and ok_sales == len(sales_rows)
        and ok_product == len(product_rows)
    )

    errors: list[SalesPlanImportErrorEntry] = []
    seen_link_names: set[str] = set()
    for row in sales_rows:
        if row.status != STATUS_LINK_NOT_FOUND:
            continue
        if row.sales_item_name in seen_link_names:
            continue
        seen_link_names.add(row.sales_item_name)
        errors.append(
            SalesPlanImportErrorEntry(
                error_code=STATUS_LINK_NOT_FOUND,
                sales_item_name=row.sales_item_name,
                target_table="tr_sales_link_name",
                target_key=f"sales_item_name={row.sales_item_name!r}",
                message="販売商品名が tr_sales_link_name に未登録です",
            )
        )
    for row in product_rows:
        if row.status != STATUS_BOM_NOT_FOUND:
            continue
        item_no = row.item_no
        errors.append(
            SalesPlanImportErrorEntry(
                error_code=STATUS_BOM_NOT_FOUND,
                sales_item_name=row.sales_item_name,
                item_no=item_no,
                excel_column=row.column,
                qty=row.qty,
                target_table="tr_item_bom",
                target_key=f"parent_item_no={item_no}" if item_no is not None else "",
                message=(
                    f"商品NO {item_no} の商品原料対照表（tr_item_bom）が未登録です"
                    if item_no is not None
                    else "商品原料対照表が未登録です"
                ),
            )
        )

    return SalesPlanImportPreviewResponse(
        year=year,
        month=month,
        file_name=file_name,
        sales_rows=sales_rows,
        product_rows=product_rows,
        errors=errors,
        summary=SalesPlanImportSummary(
            total_sales_rows=len(sales_rows),
            ok_sales_rows=ok_sales,
            total_product_rows=len(product_rows),
            ok_product_rows=ok_product,
            link_not_found=link_not_found,
            bom_not_found=bom_not_found,
            bom_missing_items=bom_missing_labels,
            merged_sales_rows=merged_sales_rows,
            duplicate_item_nos=duplicate_item_nos,
            can_register=can_register,
        ),
    )


def aggregate_sales_plan_rows(
    year: int,
    month: int,
    sales_rows: list[SalesPlanImportSalesRow],
) -> list[TeMonthlySalesPlan]:
    """同一商品NOの販売計画行を合算する（複数販売商品名→同一 item_no 対応）。"""
    merged: dict[int, TeMonthlySalesPlan] = {}
    for row in sales_rows:
        if row.status != STATUS_OK or row.item_no is None:
            continue
        item_no = int(row.item_no)
        existing = merged.get(item_no)
        if existing is None:
            merged[item_no] = TeMonthlySalesPlan(
                year=year,
                month=month,
                item_no=item_no,
                item_name=row.item_name,
                sales_size=row.sales_size,
                remarks=None,
            )
            continue
        existing.sales_size += row.sales_size
        if not existing.item_name.strip() and row.item_name.strip():
            existing.item_name = row.item_name
    return list(merged.values())


def register_from_excel(
    session: Session,
    *,
    file_name: str,
    content: bytes,
) -> SalesPlanImportRegisterResponse:
    preview = build_preview(session, file_name=file_name, content=content)
    if not preview.summary.can_register:
        raise SalesPlanExcelImportError("登録できない行があります。プレビューでエラーを確認してください。")

    year = preview.year
    month = preview.month

    TeMonthlySalesPlanRepository.delete_by_year_month_no_commit(session, year, month)
    TeMonthlyProductPlanRepository.delete_by_year_month(session, year, month)

    sales_plan_rows = aggregate_sales_plan_rows(year, month, preview.sales_rows)
    for row in sales_plan_rows:
        session.add(row)

    merged: dict[tuple[int, int, int, int], TeMonthlyProductPlan] = {}
    for row in preview.product_rows:
        key = (year, month, int(row.item_no), int(row.bulk_no))
        current = merged.get(key)
        if current is None:
            merged[key] = TeMonthlyProductPlan(
                year=year,
                month=month,
                item_no=int(row.item_no),
                bulk_no=int(row.bulk_no),
                sales_size=row.qty,
                item_name=row.item_name,
                package_size=row.package_size,
                need_size=row.need_size,
            )
        else:
            current.sales_size += row.qty
            current.need_size += row.need_size

    for product_row in merged.values():
        session.add(product_row)

    session.commit()

    return SalesPlanImportRegisterResponse(
        ok=True,
        year=year,
        month=month,
        sales_count=len(sales_plan_rows),
        product_count=len(merged),
    )
